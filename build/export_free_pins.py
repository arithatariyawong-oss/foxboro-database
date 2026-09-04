# -*- coding: utf-8 -*-
"""Every block OUTPUT pin that nothing reads, as an Excel workbook.

The question: which pins does a block have that carry no link out of it --
a value the block computes and keeps to itself. Two sources answer it
together and neither can alone:

  * `block_params.json` -- B0193AX's parameter tables, which say what pins a
    TYPE has and whether each is connectable. A pin that is not `con` cannot
    be wired at all and is not a finding; it is left out.
  * `graph.js` -- every wire in the plant, so a pin is "read" exactly when
    some edge leaves it. That file already resolved each reference against
    the right CP, including the bit-qualified ones and the Jove and ECB
    edges, so a pin feeding only the historian counts as read -- which it is.

Scope, chosen deliberately: OUTPUT pins only. The same sweep over inputs adds
427,142 more rows and almost none of them mean anything -- an AIN has 56
connectable inputs because they are settable, not because 56 wires were
expected. Outputs are the side where "nothing reads this" is a real finding.

WHAT IT CANNOT SEE, and the workbook says so on its own Notes sheet:

  * 4,607 blocks whose TYPE has no table in `block_params.json` (LONG 3,207,
    UNIVFF 893, AI 379, AO 69, STA 41, PIDFF 14, MAO 2, COMPND 2). No pin
    list means nothing to check, so they are absent rather than reported as
    clean. JOVE objects (18,502) and ECB modules (1,947) are absent on
    purpose -- neither is a Foxboro block with a B0193AX pin table.
  * a sequence (IND) block reaching a parameter from HLBL code shows up only
    because `export_sequence.py`'s edges went into graph.js... which they did
    NOT: sequence.js holds those. So an output read only by a sequence
    program can still appear here. Cross-check anything IND-adjacent against
    Sequence View before deleting it.

Writes to the FOXBORO folder, NOT into 03 WEB: it is a ~25 MB binary that has
no business in the repo.
"""
import base64
import gzip
import io
import json
import os
import time
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
WEB = os.path.dirname(HERE)
ROOT = os.path.dirname(WEB)
OUT = os.path.join(ROOT, "block-free-output-pins.xlsx")

t0 = time.time()


def load(path):
    raw = io.open(path, encoding="utf8").read()
    return json.loads(gzip.decompress(base64.b64decode(
        raw.split('"', 1)[1].rsplit('"', 1)[0])).decode("utf8"))


g = load(os.path.join(WEB, "graph.js"))
BP = json.load(io.open(os.path.join(WEB, "block_params.json"), encoding="utf8"))
nodes, edges = g["nodes"], g["edges"]
print("graph.js %d nodes, %d edges  (%.1fs)" % (len(nodes), len(edges), time.time() - t0))

# ---- what each node drives, and what drives it --------------------------
drives = defaultdict(set)          # node -> {output pin that has a reader}
fed = defaultdict(set)             # node -> {input pin that has a source}
for e in edges:
    drives[e[0]].add(e[1])
    fed[e[2]].add(e[3])

# ---- the connectable OUTPUT pins of every type --------------------------
# `OUTPUTS` on RIN and RINR is a section heading the PDF extraction picked up
# as a parameter; it is not a pin and would otherwise add one bogus row per
# RIN block in the plant.
ARTEFACT = {("RIN", "OUTPUTS"), ("RINR", "OUTPUTS")}

outs, pdesc = {}, {}
for t, b in BP.items():
    outs[t] = [p["name"] for p in b["params"]
               if p["section"] == "OUTPUTS" and p["con"]
               and (t, p["name"]) not in ARTEFACT]
    for p in b["params"]:
        if p["desc"]:
            pdesc[(t, p["name"])] = p["desc"]

# ---- how often each pin is wired ACROSS the whole plant ------------------
# The single most useful thing in the file. B0193AX lists an AIN's outputs
# alphabetically -- BAD, CRIT, HAI, HHAIND, ... -- so pin ORDER says nothing
# about which matters, and a raw list of unwired pins is dominated by alarm
# and status indicators that nobody wires anywhere, ever. Counting how many
# SIBLING blocks of the same type do wire the pin separates the two without
# anyone having to judge which pin names look important:
#   AIN.HAI  wired on 0 of 6,000 AIN blocks  -> structural, ignore it
#   AIN.PNT  wired on 5,900 of 6,000         -> this block is the odd one out
type_total = Counter()
pin_wired = Counter()
for i, n in enumerate(nodes):
    ty = n[1]
    if ty not in outs:
        continue
    type_total[ty] += 1
    for p in drives.get(i, ()):
        pin_wired[(ty, p)] += 1

# ---- walk the plant -----------------------------------------------------
rows = []
by_type, by_cp, by_area = Counter(), Counter(), Counter()
blk_type, blk_cp = Counter(), Counter()
skipped = Counter()
n_dead_block = 0

for i, n in enumerate(nodes):
    name, ty, ds, cp, area = n[0], n[1], n[2], n[3], n[4]
    if ty not in outs:
        skipped[ty] += 1
        continue
    pins = outs[ty]
    if not pins:
        continue
    wired = drives.get(i, ())
    free = [p for p in pins if p not in wired]
    if not free:
        continue
    comp, _, blk = name.partition(":")
    if not blk:
        comp, blk = "", name
    live = bool(wired) or bool(fed.get(i))
    all_dead = not wired                       # nothing reads ANY of its outputs
    if all_dead:
        n_dead_block += 1
    blk_type[ty] += 1
    blk_cp[cp] += 1
    for p in free:
        sib = pin_wired.get((ty, p), 0)
        tot = type_total.get(ty, 0)
        rows.append((cp, comp, blk, ty, ds, area, p, pdesc.get((ty, p), ""),
                     sib, tot, round(sib / tot, 4) if tot else 0,
                     len(wired), len(pins),
                     "ใช่" if all_dead else "ไม่",
                     "ใช่" if live else "ไม่"))
        by_type[ty] += 1
        by_cp[cp] += 1
        by_area[area or "(ว่าง)"] += 1

rows.sort(key=lambda r: (r[0], r[1], r[2], r[6]))
print("free output pins: %d over %d blocks  (%.1fs)"
      % (len(rows), sum(blk_type.values()), time.time() - t0))
print("blocks whose outputs are read by NOBODY: %d" % n_dead_block)

# ---- the workbook -------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="26505C")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)

wb = Workbook(write_only=True)

COLS = [("CP", 11), ("COMPOUND", 17), ("BLOCK", 20), ("TYPE", 9),
        ("DESCRP", 34), ("AREA", 13), ("PIN", 11),
        ("คำอธิบายขา (B0193AX)", 40),
        ("บล็อก TYPE เดียวกันที่ต่อขานี้", 13),
        ("บล็อก TYPE นี้ทั้งหมด", 12),
        ("สัดส่วนที่ต่อ", 11),
        ("ขาออกที่ต่อแล้ว", 9), ("ขาออกทั้งหมด", 9),
        ("ทั้งบล็อกไม่มีใครอ่านเลย", 11), ("บล็อกนี้มีสายอื่นอยู่", 11)]

# ---- the findings, up front ---------------------------------------------
# 80% of the rows sit at a ratio under 0.02 -- pins no block of that type
# wires anywhere. Sorting the other end brings the actual anomalies to the
# top: 764 of 765 MOVLV blocks wire COUT_1 and CBM_A:32HV001 does not; four
# PIDA blocks compute an OUT that nothing reads. That is a list an engineer
# can work through, so it gets its own sheet rather than waiting to be found
# by someone who thinks to sort a 487,364-row table by column K.
SUSPECT = 0.5
sus = sorted((r for r in rows if r[10] >= SUSPECT), key=lambda r: -r[10])
s = wb.create_sheet("ขาที่น่าสงสัย")
s.freeze_panes = "A2"
for k, (h, w) in enumerate(COLS, 1):
    s.column_dimensions[get_column_letter(k)].width = w
if sus:
    s.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(COLS)), len(sus) + 1)
hh = []
for h, _ in COLS:
    c = WriteOnlyCell(s, value=h)
    c.fill, c.font = HEAD_FILL, HEAD_FONT
    c.alignment = Alignment(vertical="center", wrap_text=True)
    hh.append(c)
s.append(hh)
for r in sus:
    s.append(list(r))
print("suspicious (ratio >= %.2f): %d rows" % (SUSPECT, len(sus)))



ws = wb.create_sheet("ขาออกไม่มีคนอ่าน")
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(COLS)), len(rows) + 1)
for k, (h, w) in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(k)].width = w
hdr = []
for h, _ in COLS:
    c = WriteOnlyCell(ws, value=h)
    c.fill, c.font = HEAD_FILL, HEAD_FONT
    c.alignment = Alignment(vertical="center", wrap_text=True)
    hdr.append(c)
ws.append(hdr)
for r in rows:
    ws.append(list(r))
print("sheet 1 written  (%.1fs)" % (time.time() - t0))


def summary(title, counter, label, extra=None):
    s = wb.create_sheet(title)
    s.freeze_panes = "A2"
    cols = [(label, 26), ("ขาออกที่ไม่มีคนอ่าน", 18)] + (extra or [])
    for k, (h, w) in enumerate(cols, 1):
        s.column_dimensions[get_column_letter(k)].width = w
    hh = []
    for h, _ in cols:
        c = WriteOnlyCell(s, value=h)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        hh.append(c)
    s.append(hh)
    return s


s = summary("สรุปตาม TYPE", by_type, "TYPE", [("บล็อกที่มีขาว่าง", 16)])
for t, c in by_type.most_common():
    s.append([t, c, blk_type[t]])

s = summary("สรุปตาม CP", by_cp, "CP NAME", [("บล็อกที่มีขาว่าง", 16)])
for t, c in by_cp.most_common():
    s.append([t, c, blk_cp[t]])

s = summary("สรุปตาม AREA", by_area, "AREA")
for t, c in by_area.most_common():
    s.append([t, c])

# ---- the notes sheet: what this does and does not cover -----------------
note = wb.create_sheet("หมายเหตุ")
note.column_dimensions["A"].width = 118
LINES = [
    ("ขาออกของบล็อกที่ไม่มีใครอ่าน — Foxboro I/A Series", True),
    ("", False),
    ("สร้างโดย 03 WEB/build/export_free_pins.py เมื่อ %s"
     % time.strftime("%Y-%m-%d %H:%M"), False),
    ("", False),
    ("นับยังไง", True),
    ("• รายชื่อขาของแต่ละ TYPE มาจาก block_params.json (ตารางพารามิเตอร์ของ B0193AX)", False),
    ("  เอาเฉพาะขาในหมวด OUTPUTS ที่ต่อสายได้จริง (con = true) ขาที่ต่อไม่ได้ไม่นับ", False),
    ("• ถือว่า 'มีคนอ่าน' เมื่อมี edge ออกจากขานั้นใน graph.js ซึ่งรวมสายพารามิเตอร์", False),
    ("  สายฮาร์ดแวร์ (ECB/FBM) และสาย Jove แล้ว — ขาที่ส่งให้ historian อย่างเดียวจึงนับว่าอ่านแล้ว", False),
    ("• ขาเข้าไม่ได้อยู่ในไฟล์นี้ (อีก 427,142 แถว) เพราะ AIN มีขาเข้าต่อได้ 56 ขา", False),
    ("  ด้วยเหตุผลว่าตั้งค่าได้ ไม่ใช่ว่าต้องมี 56 สาย — ฝั่ง output ต่างหากที่ 'ไม่มีใครอ่าน' มีความหมาย", False),
    ("", False),
    ("สิ่งที่ไฟล์นี้มองไม่เห็น — อ่านก่อนใช้ตัดสินใจลบอะไร", True),
    ("• บล็อก 4,607 ตัวที่ TYPE ไม่มีตารางขาใน block_params.json จะไม่อยู่ในไฟล์นี้เลย", False),
    ("  (LONG 3,207 · UNIVFF 893 · AI 379 · AO 69 · STA 41 · PIDFF 14 · MAO 2 · COMPND 2)", False),
    ("  ไม่มีรายการขา = ไม่มีอะไรให้ตรวจ จึงหายไปเฉย ๆ ไม่ใช่ว่าสะอาด", False),
    ("• JOVE object (18,502) และ ECB module (1,947) ไม่อยู่ในไฟล์นี้โดยตั้งใจ", False),
    ("  ทั้งสองอย่างไม่ใช่บล็อก Foxboro ที่มีตารางขาแบบ B0193AX", False),
    ("• **บล็อก sequence (IND) ที่อ่านค่าจากโค้ด HLBL ไม่ได้อยู่ใน graph.js**", False),
    ("  สายพวกนั้นอยู่ใน sequence.js ต่างหาก ขาที่ถูกอ่านโดยโปรแกรม sequence เท่านั้น", False),
    ("  จึงยังโผล่มาในไฟล์นี้ได้ — ก่อนลบอะไรที่เกี่ยวกับ IND ให้เช็คใน Sequence View ก่อน", False),
    ("", False),
    ("คอลัมน์ที่ควรใช้ก่อน — 'สัดส่วนที่ต่อ'", True),
    ("B0193AX เรียงขาออกตามตัวอักษร (AIN คือ BAD, CRIT, HAI, HHAIND, … , PNT)", False),
    ("ลำดับขาจึงไม่ได้บอกว่าขาไหนสำคัญ และรายการขาว่างดิบ ๆ จะเต็มไปด้วยขาแจ้ง", False),
    ("สถานะ/อะลาร์มที่ไม่มีใครต่อที่ไหนเลยอยู่แล้ว  จึงเพิ่มคอลัมน์นี้ให้แยกออกจากกัน", False),
    ("โดยไม่ต้องเดาว่าชื่อขาไหนดูสำคัญ:", False),
    ("• สัดส่วนที่ต่อ ≈ 0  → บล็อกชนิดเดียวกันทั้งโรงงานก็ไม่ต่อขานี้ (เช่น AIN.HAI)", False),
    ("  เป็นเรื่องปกติของชนิดบล็อก ไม่ใช่สิ่งผิดปกติ", False),
    ("• สัดส่วนที่ต่อ สูง  → พี่น้องชนิดเดียวกันเขาต่อกันเกือบหมด แต่บล็อกนี้ไม่ต่อ", False),
    ("  ← **เรียงคอลัมน์นี้จากมากไปน้อยก่อน แล้วดูจากบนลงมา**", False),
    ("", False),
    ("คอลัมน์อื่น", True),
    ("• ขาออกที่ต่อแล้ว / ขาออกทั้งหมด — ดูว่าบล็อกนี้ใช้ขาไปกี่ขาจากที่มี", False),
    ("• ทั้งบล็อกไม่มีใครอ่านเลย = ใช่ → ไม่มีขาออกสักขาที่มีคนอ่าน (บล็อกคำนวณทิ้ง)", False),
    ("• บล็อกนี้มีสายอื่นอยู่ = ไม่ → บล็อกนี้ไม่มีสายอะไรเลยทั้งเข้าและออก", False),
    ("", False),
    ("ตัวเลขรวม", True),
    ("• ขาออกที่ไม่มีคนอ่าน: %s แถว" % format(len(rows), ","), False),
    ("• บล็อกที่มีขาออกว่างอย่างน้อยหนึ่งขา: %s" % format(sum(blk_type.values()), ","), False),
    ("• บล็อกที่ไม่มีขาออกไหนถูกอ่านเลย: %s" % format(n_dead_block, ","), False),
]
for text, bold in LINES:
    c = WriteOnlyCell(note, value=text)
    if bold:
        c.font = Font(bold=True, size=11, color="26505C")
    note.append([c])

wb.save(OUT)
print("wrote %s  (%.1f MB, %.1fs)"
      % (OUT, os.path.getsize(OUT) / 1e6, time.time() - t0))
