# -*- coding: utf-8 -*-
"""Build the per-system hardware inventory for SYSTEM MONITOR and write systems.js.

Three sources meet here:

  * data.js -- the tag database. Every ECB block is a piece of hardware: the
    module ECBs (ECB1/2/5/12/200/202/...) carry DEV_ID, the letterbug printed
    on the FBM, and HWTYPE, the module's model code. The child ECBs (ECB201
    for a HART/FF/serial device, ECB18 for an FBM43 transmitter) carry PARENT
    = their module's letterbug and name the channel they sit on -- ECB201 in
    DVNAME ("CH3"), ECB18 in CHAN ("3").
    An I/O block reaches its hardware by IOM_ID. For a channel-addressed FBM
    that is the module itself and PNT_NO is the channel number; for a HART or
    FF module it is the *child* ECB, and PNT_NO is a variable name ("CURRENT"),
    so the channel has to be read off the child. Both paths are followed here.

  * TOP-Foxboro-Hardware-2025_RevA-1.xlsx -- the plant's own hardware register.
    `FBMnameAndType` gives AI/AO/DI/DO/PI/IT/FF counts for 1,062 letterbugs,
    which is what a spare count has to be measured against; `DATAFBMlist` gives
    each letterbug's model, firmware rev and the rev checker's warning; and
    `FBMandCP` gives every model's product number, description and lifecycle
    phase, plus how many are installed and how many sit on the shelf.

  * the observed maximum channel. FBM07 and FBM09 accept an expander (FBM12,
    FBM42) that doubles the point count on the SAME letterbug, so a module can
    legitimately carry channel 27 while the register says 16. Capacity is
    therefore max(register, observed) and the module is flagged `exp` when the
    observed count wins, rather than reporting a negative spare.

Spare is only counted where a channel actually exists. A serial gateway
(FBM230/231/232/233) addresses Modbus registers, not channels, and an FF
module addresses devices on a segment; neither has a fixed point count, so
both are reported with their point/device count and no spare figure. Saying
"0 spare" there would be a lie, and saying "16 spare" a worse one.

Output systems.js mirrors data.js and graph.js: gzip + base64 in a global,
inflated by the page with DecompressionStream, because a page opened from
file:// cannot fetch() a .json.
"""
import base64, gzip, json, os, sys, time
from collections import defaultdict, Counter

HERE = os.path.dirname(os.path.abspath(__file__))
WEB = os.path.dirname(HERE)
ROOT = os.path.dirname(WEB)
DATA = os.path.join(WEB, "data.js")
XLSX = os.path.join(ROOT, "TOP-Foxboro-Hardware-2025_RevA-1.xlsx")
OUT = os.path.join(WEB, "systems.js")

t0 = time.time()

# ---------------------------------------------------------------- data.js --
raw = open(DATA, encoding="utf8").read()
b64 = raw.split('"', 1)[1].rsplit('"', 1)[0]
d = json.loads(gzip.decompress(base64.b64decode(b64)).decode("utf8"))
N, H, C = d["n"], d["h"], d["c"]
IDX = {h: i for i, h in enumerate(H) if h}


def dense(col):
    a = [-1] * N
    sp = C[IDX[col]]
    if not isinstance(sp, dict):
        return a, []
    v = sp["v"]
    if "r" in sp:
        row = 0
        for k, dl in enumerate(sp["r"]):
            row += dl
            a[row] = v[k]
    else:
        for k, c in enumerate(v):
            a[k] = c
    return a, sp["d"]


def col(name):
    a, dd = dense(name)
    return [dd[x] if x >= 0 else "" for x in a]


TY = col("TYPE"); CP = col("CP NAME"); NM = col("NAME"); DS = col("DESCRP")
DEV = col("DEV_ID"); HW = col("HWTYPE"); IOM = col("IOM_ID"); IOMR = col("IOMIDR")
PNT = col("PNT_NO"); PAR = col("PARENT"); DVN = col("DVNAME"); CHN = col("CHAN")
PRT = col("PORTNO"); AR = col("AREA"); PER = col("PERIOD")
print("data.js %d rows (%.1fs)" % (N, time.time() - t0))

# --------------------------------------------------------- hardware register
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

# model -> product number, description, lifecycle, shelf spares, installed
MODEL = {}
for r in wb["FBMandCP"].iter_rows(min_row=4, values_only=True):
    if not r[0]:
        continue
    m = str(r[0]).strip()
    prev = MODEL.get(m)
    rec = {
        "pn": str(r[2] or ""), "md": str(r[3] or ""), "lc": str(r[4] or ""),
        "shelf": int(r[5] or 0), "inst": int(r[6] or 0), "cls": str(r[1] or ""),
    }
    # FBM205 is listed twice (two product numbers); keep the larger install
    if prev is None or rec["inst"] >= prev["inst"]:
        MODEL[m] = rec

# (cp, letterbug) -> model, [AI, AO, DI, DO, PI, IT, FF]
CLS = ["AI", "AO", "DI", "DO", "PI", "IT", "FF"]
CAPS = {}
for r in wb["FBMnameAndType"].iter_rows(min_row=2, values_only=True):
    if not r or not r[1] or not r[2]:
        continue
    CAPS[(str(r[1]).strip(), str(r[2]).strip())] = (
        str(r[3] or "").strip(), [int(x or 0) for x in r[4:11]])

# letterbug -> model, description, sw rev, latest rev, hw rev, warning
FW = {}
for r in wb["DATAFBMlist"].iter_rows(min_row=2, values_only=True):
    if not r or not r[0] or str(r[0]) == "Device":
        continue
    FW[str(r[0]).strip()] = {
        "m": str(r[1] or "").strip(), "md": str(r[2] or "").strip(),
        "sw": str(r[3] or "").strip(), "latest": str(r[5] or "").strip(),
        "hwr": str(r[7] or "").strip(),
        "warn": "" if r[9] in (None, "") else str(r[9]).strip(),
    }
print("register: %d models, %d letterbug capacities, %d firmware rows"
      % (len(MODEL), len(CAPS), len(FW)))

# ---------------------------------------------------- module type reference
# hwtype -> (model, kind, nominal [AI,AO,DI,DO,PI,IT,FF])
#   chan = PNT_NO is the channel number
#   hart = channel lives on an ECB201 child, named in DVNAME as "CHn"
#   itx  = FBM43 intelligent transmitter; channel on an ECB18 child, CHAN
#   ff   = FOUNDATION fieldbus segment; devices, not channels
#   comm = serial/ethernet gateway; Modbus registers, not channels
Z = [0] * 7


def cap(ai=0, ao=0, di=0, do=0, pi=0, it=0, ff=0):
    return [ai, ao, di, do, pi, it, ff]


HWREF = {
    "1":   ("FBM01",   "chan", cap(ai=8)),
    "2":   ("FBM02",   "chan", cap(ai=8)),
    "5":   ("FBM05",   "chan", cap(ai=4, ao=4)),
    "6":   ("FBM06",   "chan", cap(ai=4, ao=4)),
    "7":   ("FBM07",   "chan", cap(di=16)),
    "9":   ("FBM09",   "chan", cap(di=8, do=8)),
    "32":  ("FBM32",   "other", Z),
    "41":  ("FBM41",   "chan", cap(di=8, do=8)),
    "43":  ("FBM43",   "itx",  cap(ai=8)),
    "200": ("FBM200",  "other", Z),
    "201": ("FBM201",  "chan", cap(ai=8)),
    "202": ("FBM202",  "chan", cap(ai=8)),
    "204": ("FBM204",  "chan", cap(ai=4, ao=4)),
    "205": ("FBM205",  "chan", cap(ai=4, ao=4)),
    "206": ("FBM206",  "chan", cap(pi=8)),
    "207": ("FBM207B", "chan", cap(di=16)),
    "214": ("FBM214B", "hart", cap(ai=8)),
    "216": ("FBM216B", "hart", cap(ai=8)),
    "217": ("FBM217",  "chan", cap(di=32)),
    "218": ("FBM218",  "hart", cap(ao=8)),
    "228": ("FB228R",  "ff",   cap(ff=1)),
    "230": ("FBM230",  "comm", Z),
    "231": ("FBM231",  "comm", Z),
    "232": ("FBM232",  "comm", Z),
    "233": ("FBM233",  "comm", Z),
    "240": ("FBM240",  "chan", cap(do=16)),
    "241": ("FBM241C", "chan", cap(di=8, do=8)),
    "242": ("FBM242",  "chan", cap(do=16)),
    "245": ("FBM245",  "hart", cap(ai=4, ao=4)),
}

# The order the classes occupy the channel numbers, verified against the
# database: FBM05/205/245 run AI 1-4 then AO 5-8; FBM09/41/241 run DI 1-8 then
# DO 9-16; the single-class modules run 1-n.
ORDER = ["AI", "AO", "PI", "DI", "DO", "IT", "FF"]

# No 200 Series FBM carries more than 32 points, and an expander at most
# doubles that. Anything above this is a mapped register, not a channel.
CH_MAX = 64

MODULE_ECB = {"ECB1", "ECB2", "ECB4", "ECB5", "ECB7", "ECB12", "ECB16",
              "ECB110", "ECB200", "ECB202"}
CHILD_ECB = {"ECB201", "ECB18"}

INBLK = {"AIN", "AINR", "CIN", "CINR", "RIN", "BIN", "IIN", "PAKIN", "MAIN",
         "MCIN", "STRIN", "AI", "UNIVFF", "PIDFF", "MRIN"}

# ------------------------------------------------------------- index the DB
ecb_row = {}                                  # (cp, dev) -> row
for r in range(N):
    if TY[r].startswith("ECB") and DEV[r]:
        ecb_row.setdefault((CP[r], DEV[r]), r)

children = defaultdict(list)                  # (cp, parent dev) -> child rows
for r in range(N):
    if TY[r] in CHILD_ECB and PAR[r]:
        children[(CP[r], PAR[r])].append(r)

# every I/O block, filed under the ECB it names
bound = defaultdict(list)                     # (cp, dev) -> block rows
for r in range(N):
    if TY[r].startswith("ECB"):
        continue
    for dev in (IOM[r], IOMR[r]):
        if dev:
            bound[(CP[r], dev)].append(r)

# ------------------------------------------------------------- the systems
db_names = sorted(set(CP[r] for r in range(N) if CP[r]))
# 02CP02 and 03CP10 carry FBMs in the register but were not in the SaveAll
# export; list them so the module inventory is not silently short of them.
reg_names = sorted({str(r[1]).strip()
                    for r in wb["FBMnameAndType"].iter_rows(min_row=2, values_only=True)
                    if r and r[1]} - set(db_names))
sys_names = sorted(db_names + reg_names)
sys_idx = {c: i for i, c in enumerate(sys_names)}
DB_SYS = set(db_names)

STATION = {}
for r in wb["CP"].iter_rows(values_only=True):
    if r and r[0]:
        STATION[str(r[0]).strip()] = str(r[1] or "").strip()

systems = []
for c in sys_names:
    rows = [r for r in range(N) if CP[r] == c]
    areas = Counter(AR[r] for r in rows if AR[r])
    st = STATION.get(c, "")
    fw = FW.get(c, {})
    systems.append({
        "n": c,
        "sta": st or fw.get("m", ""),
        "std": fw.get("md", ""),
        "area": areas.most_common(1)[0][0] if areas else "",
        "blocks": len(rows),
        "comps": len(set(NM[r].split(":")[0] for r in rows if NM[r])),
        "types": dict(Counter(TY[r] for r in rows if TY[r]).most_common()),
        "per": dict(sorted(Counter(PER[r] for r in rows if PER[r]).items(),
                           key=lambda kv: int(kv[0]) if kv[0].isdigit() else 99)),
        "mods": [],
        "unmapped": [],
        "src": "db" if c in DB_SYS else "hw",
    })

# --------------------------------------------------------------- modules --
mods = []
seen_dev = set()

for r in range(N):
    if TY[r] not in MODULE_ECB or not DEV[r]:
        continue
    c, dev = CP[r], DEV[r]
    if (c, dev) in seen_dev:
        continue
    seen_dev.add((c, dev))
    hwt = HW[r]
    model, kind, nominal = HWREF.get(hwt, ("", "other", Z))

    # the plant register overrides the type default, letterbug by letterbug
    reg = CAPS.get((c, dev))
    if reg:
        if reg[0]:
            model = reg[0]
            kind = HWREF.get(hwt, (None, kind, None))[1]
        if any(reg[1]):
            nominal = reg[1]
    fw = FW.get(dev, {})
    if fw.get("m"):
        model = fw["m"]
    mrec = MODEL.get(model, {})

    kids = children.get((c, dev), [])
    blocks = list(bound.get((c, dev), []))

    # ---- lay the channels out ------------------------------------------
    slots = []                                # (class, channel number)
    for cl in ORDER:
        n = nominal[CLS.index(cl)] if cl in CLS else 0
        if kind == "ff":
            n = 0
        for _ in range(n):
            slots.append(cl)
    ch = []                                   # [class, tag, type, descrp, dev]
    over = False
    pts = 0

    if kind in ("chan",):
        # the block's own PNT_NO is the channel. A value past CH_MAX is a
        # register address on a module that also carries mapped data, not a
        # physical channel -- count the block, but never stretch the module.
        occupied = defaultdict(list)
        for b in blocks:
            if PNT[b].isdigit() and int(PNT[b]) <= CH_MAX:
                occupied[int(PNT[b])].append(b)
            else:
                pts += 1
        top = max([len(slots)] + list(occupied)) if (slots or occupied) else 0
        if occupied and max(occupied) > len(slots):
            over = True                        # an expander is fitted
        for i in range(1, top + 1):
            cl = slots[i - 1] if i <= len(slots) else (slots[-1] if slots else "")
            b = occupied.get(i)
            if b:
                pts += len(b)
                b = b[0]
                ch.append([cl, NM[b], TY[b], DS[b], ""])
            else:
                ch.append([cl, "", "", "", ""])

    elif kind in ("hart", "itx"):
        # The channel is named on the child ECB -- ECB201 writes it in DVNAME
        # as "CH3", ECB18 in CHAN as "3". A handful of ECB201s carry the
        # device letterbug in DVNAME instead ("1092A4"); those devices still
        # occupy a channel, so they take the lowest free one rather than being
        # dropped, which would under-report the module as spare.
        by_ch, loose = {}, []
        for k in kids:
            lab = (DVN[k] if TY[k] == "ECB201" else CHN[k]).strip()
            n = None
            if lab[:2].upper() == "CH" and lab[2:].isdigit():
                n = int(lab[2:])
            elif lab.isdigit() and 1 <= int(lab) <= max(len(slots), 1):
                n = int(lab)
            if n is None or n in by_ch or n > CH_MAX:
                loose.append(k)
            else:
                by_ch[n] = k
        for k in loose:
            n = next(i for i in range(1, CH_MAX + 2) if i not in by_ch)
            by_ch[n] = k
        top = max([len(slots)] + list(by_ch)) if (slots or by_ch) else 0
        if by_ch and max(by_ch) > len(slots):
            over = True
        for i in range(1, top + 1):
            cl = slots[i - 1] if i <= len(slots) else (slots[-1] if slots else "")
            k = by_ch.get(i)
            if k is None:
                ch.append([cl, "", "", "", ""])
                continue
            kb = bound.get((c, DEV[k]), [])
            pts += len(kb)
            if kb:
                b = kb[0]
                ch.append([cl, NM[b], TY[b], DS[b], DEV[k]])
            else:
                # the device ECB is configured but no block reads it
                ch.append([cl, "", "", "", DEV[k]])
        pts += len(blocks)

    else:
        # ff / comm / other: count what is attached, claim no channel map
        pts = len(blocks) + sum(len(bound.get((c, DEV[k]), [])) for k in kids)

    used = sum(1 for x in ch if x[1])
    held = sum(1 for x in ch if not x[1] and x[4])    # ECB present, no tag
    spare = len(ch) - used - held

    mods.append({
        "s": sys_idx[c],
        "d": dev,
        "e": TY[r],
        "b": NM[r],
        "h": hwt,
        "m": model,
        "md": (mrec.get("md") or fw.get("md") or ""),
        "pn": mrec.get("pn", ""),
        "lc": mrec.get("lc", ""),
        "k": kind,
        "cap": nominal,
        "ch": ch,
        "u": used,
        "hd": held,
        "sp": spare,
        "pts": pts,
        "kids": len(kids),
        "sw": fw.get("sw", ""),
        "lt": fw.get("latest", ""),
        "hr": fw.get("hwr", ""),
        "wn": fw.get("warn", ""),
        "ar": AR[r],
        "row": r,
        "exp": 1 if over else 0,
        "src": "db+hw" if (dev in FW or reg) else "db",
    })

for i, m in enumerate(mods):
    systems[m["s"]]["mods"].append(i)

# ---- letterbugs the register knows and the tag database does not ---------
# the plant reuses letterbugs across stations, so the pair is the key
db_dev = {(systems[m_["s"]]["n"], m_["d"]) for m_ in mods}
for (c, dev), (model, capv) in sorted(CAPS.items()):
    if (c, dev) in db_dev or c not in sys_idx:
        continue
    fw = FW.get(dev, {})
    mrec = MODEL.get(model, {})
    hwt = next((k for k, v in HWREF.items() if v[0] == model), "")
    kind = HWREF.get(hwt, ("", "other", Z))[1]
    slots = []
    for cl in ORDER:
        n = capv[CLS.index(cl)] if cl in CLS else 0
        if kind == "ff":
            n = 0
        slots += [cl] * n
    mods.append({
        "s": sys_idx[c], "d": dev, "e": "", "b": "", "h": hwt, "m": model,
        "md": (mrec.get("md") or fw.get("md") or ""), "pn": mrec.get("pn", ""),
        "lc": mrec.get("lc", ""), "k": kind, "cap": capv,
        "ch": [[cl, "", "", "", ""] for cl in slots],
        "u": 0, "hd": 0, "sp": len(slots), "pts": 0, "kids": 0,
        "sw": fw.get("sw", ""), "lt": fw.get("latest", ""), "hr": fw.get("hwr", ""),
        "wn": fw.get("warn", ""), "ar": "", "row": -1, "exp": 0, "src": "hw",
    })
    systems[sys_idx[c]]["mods"].append(len(mods) - 1)

# ---- I/O bindings that reach no ECB at all ------------------------------
for c in sys_names:
    miss = Counter()
    for r in range(N):
        if CP[r] != c or TY[r].startswith("ECB") or not IOM[r]:
            continue
        if (c, IOM[r]) not in ecb_row:
            miss[IOM[r]] += 1
    if miss:
        systems[sys_idx[c]]["unmapped"] = miss.most_common(40)

# ------------------------------------------------------------------ write --
payload = {
    "gen": time.strftime("%Y-%m-%d"),
    "cls": CLS,
    "sys": systems,
    "mods": mods,
    "models": MODEL,
}
js = json.dumps(payload, separators=(",", ":"), ensure_ascii=False).encode("utf8")
gz = gzip.compress(js, 9)
open(OUT, "w", encoding="utf8").write(
    'window.FOX_SYS_B64="%s";\n' % base64.b64encode(gz).decode("ascii"))

tot_ch = sum(len(m["ch"]) for m in mods)
tot_u = sum(m["u"] for m in mods)
tot_h = sum(m["hd"] for m in mods)
tot_s = sum(m["sp"] for m in mods)
print("systems %d, modules %d (%d from register only)"
      % (len(systems), len(mods), sum(1 for m in mods if m["src"] == "hw")))
print("channels %d: used %d, ECB-only %d, spare %d" % (tot_ch, tot_u, tot_h, tot_s))
print("json %.2f MB -> gzip %.2f MB -> systems.js %.2f MB  (%.1fs)"
      % (len(js) / 1e6, len(gz) / 1e6, os.path.getsize(OUT) / 1e6, time.time() - t0))
